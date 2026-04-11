"""
根据项目 Presets 配置生成/更新 Excel，格式与 ExcelToLua3 兼容。
- 第 1 行：可选注释（中文说明）
- 第 2 行：表头（英文字段名）
- 第 3 行：类型（string / int / float / table）
- 第 4 行起：数据

运行：在 Excels 目录下执行  python ConfigToExcel.py
"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: pip install openpyxl\n"
        f"Details: {exc}"
    )

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR  # 输出到 Excels 目录

# -------- PlayerSystem Presets.Levels (1-120) --------
PLAYER_LEVEL_UP_EXP = [
    10, 20, 30, 40, 50, 60, 70, 80, 90, 100,
    200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800, 2000,
    4000, 8000, 12000, 16000, 20000, 24000, 28000, 32000, 36000, 40000,
    80000, 160000, 240000, 320000, 400000, 480000, 560000, 640000, 720000, 800000,
    1600000, 3200000, 4800000, 6400000, 8000000, 9600000, 11200000, 12800000, 14400000, 16000000,
    32000000, 64000000, 96000000, 128000000, 160000000, 192000000, 224000000, 256000000, 288000000, 320000000,
    640000000, 1280000000, 1920000000, 2560000000, 3200000000, 3840000000, 4480000000, 5120000000, 5760000000, 6400000000,
    12800000000, 25600000000, 38400000000, 51200000000, 64000000000, 76800000000, 89600000000, 102400000000, 115200000000, 128000000000,
    256000000000, 512000000000, 768000000000, 1024000000000, 1280000000000, 1536000000000, 1792000000000, 2048000000000, 2304000000000, 2560000000000,
    5120000000000, 10240000000000, 15360000000000, 20480000000000, 25600000000000, 30720000000000, 35840000000000, 40960000000000, 46080000000000, 51200000000000,
    102400000000000, 204800000000000, 307200000000000, 409600000000000, 512000000000000, 614400000000000, 716800000000000, 819200000000000, 921600000000000, 1024000000000000,
    2048000000000000, 4096000000000000, 6144000000000000, 8192000000000000, 10240000000000000, 12288000000000000, 14336000000000000, 16384000000000000, 18432000000000000, 20480000000000000,
]

# -------- WeaponSystem Presets.Weapons（与 Keys.WeaponTypes 一致：Pickaxe/Laser/Bomb/...）--------
WEAPONS_HEADERS = [
    "weaponId", "name", "weaponType", "baseDamage", "damageRange", "coolDown", "attackRange", "levelReq",
    "RayLength", "maxBlocks", "explosionSize", "fuseTime", "drillWidth",
    "layers", "radius", "orbitSpeed", "selfRotationSpeed",
    "projectileCount", "bounceCount", "projectileSpeed",
]
WEAPONS_TYPES = [
    "string", "string", "string", "float", "float", "float", "int", "int",
    "int", "int", "int", "float", "int",
    "int", "int", "float", "float",
    "int", "int", "float",
]

WEAPONS_DATA = [
    {"weaponId": "Pickaxe1", "name": "Iron Pickaxe", "weaponType": "Pickaxe", "baseDamage": 2, "damageRange": 0.5, "coolDown": 0.7, "attackRange": 30, "levelReq": 1},
    {"weaponId": "Pickaxe2", "name": "Strong Pickaxe", "weaponType": "Pickaxe", "baseDamage": 3, "damageRange": 1, "coolDown": 0.7, "attackRange": 30, "levelReq": 3},
    {"weaponId": "Pickaxe3", "name": "Metal Pickaxe", "weaponType": "Pickaxe", "baseDamage": 4.5, "damageRange": 2, "coolDown": 0.7, "attackRange": 30, "levelReq": 5},
    {"weaponId": "Laser1", "name": "Laser", "weaponType": "Laser", "baseDamage": 3, "RayLength": 50, "attackRange": 30, "coolDown": 0.1, "maxBlocks": 3, "levelReq": 1},
    {"weaponId": "Laser2", "name": "Strong Laser", "weaponType": "Laser", "baseDamage": 4.5, "RayLength": 50, "attackRange": 30, "coolDown": 0.1, "maxBlocks": 5, "levelReq": 4},
    {"weaponId": "Laser3", "name": "Super Laser", "weaponType": "Laser", "baseDamage": 9, "RayLength": 50, "attackRange": 30, "coolDown": 0.1, "maxBlocks": 10, "levelReq": 7},
    {"weaponId": "Bomb1", "name": "Bomb", "weaponType": "Bomb", "baseDamage": 15, "attackRange": 30, "coolDown": 1.2, "explosionSize": 1, "fuseTime": 2, "levelReq": 1},
    {"weaponId": "Bomb2", "name": "Grenade", "weaponType": "Bomb", "baseDamage": 22, "attackRange": 30, "coolDown": 1.2, "explosionSize": 1, "fuseTime": 2, "levelReq": 6},
    {"weaponId": "Bomb3", "name": "TNT", "weaponType": "Bomb", "baseDamage": 33, "attackRange": 30, "coolDown": 1.2, "explosionSize": 2, "fuseTime": 2, "levelReq": 9},
    {"weaponId": "Drill1", "name": "Drill", "weaponType": "Drill", "baseDamage": 25, "attackRange": 10, "coolDown": 0.5, "drillWidth": 1, "levelReq": 15},
    {"weaponId": "Drill2", "name": "Cute Drill", "weaponType": "Drill", "baseDamage": 38, "attackRange": 10, "coolDown": 0.5, "drillWidth": 1, "levelReq": 22},
    {"weaponId": "Drill3", "name": "Powerful Drill", "weaponType": "Drill", "baseDamage": 57, "attackRange": 10, "coolDown": 0.5, "drillWidth": 2, "levelReq": 30},
    {"weaponId": "OrbitingBoomerang1", "name": "Orbiting Boomerang", "weaponType": "OrbitingBoomerang", "baseDamage": 20, "attackRange": 30, "coolDown": 1, "layers": 1, "radius": 6, "orbitSpeed": 360, "selfRotationSpeed": 720, "levelReq": 10},
    {"weaponId": "OrbitingBoomerang2", "name": "Quad Boomerang", "weaponType": "OrbitingBoomerang", "baseDamage": 30, "attackRange": 30, "coolDown": 1, "layers": 2, "radius": 9, "orbitSpeed": 500, "selfRotationSpeed": 720, "levelReq": 15},
    {"weaponId": "OrbitingBoomerang3", "name": "Cyborg Boomerang", "weaponType": "OrbitingBoomerang", "baseDamage": 45, "attackRange": 30, "coolDown": 1, "layers": 3, "radius": 12, "orbitSpeed": 700, "selfRotationSpeed": 1080, "levelReq": 20},
    {"weaponId": "BouncingBomb1", "name": "Bouncing Bomb", "weaponType": "BouncingBomb", "baseDamage": 100, "attackRange": 30, "coolDown": 1.2, "projectileCount": 5, "bounceCount": 1, "fuseTime": 3, "projectileSpeed": 40, "levelReq": 15},
    {"weaponId": "BouncingBomb2", "name": "Bouncing Fish", "weaponType": "BouncingBomb", "baseDamage": 150, "attackRange": 30, "coolDown": 1.2, "projectileCount": 5, "bounceCount": 2, "fuseTime": 3, "projectileSpeed": 45, "levelReq": 22},
    {"weaponId": "BouncingBomb3", "name": "Bouncing Bear", "weaponType": "BouncingBomb", "baseDamage": 220, "attackRange": 30, "coolDown": 1.2, "projectileCount": 5, "bounceCount": 3, "fuseTime": 3, "projectileSpeed": 50, "levelReq": 30},
]


def write_sheet(ws, comment_row, headers, type_row, data_rows):
    """写入一张表：第 1 行注释，第 2 行表头，第 3 行类型，第 4 行起数据。"""
    for col, h in enumerate(headers, start=1):
        if comment_row and col <= len(comment_row):
            ws.cell(row=1, column=col, value=comment_row[col - 1])
        ws.cell(row=2, column=col, value=h)
        ws.cell(row=3, column=col, value=type_row[col - 1] if col <= len(type_row) else None)
    for r, row in enumerate(data_rows, start=4):
        for c, val in enumerate(row, start=1):
            if c <= len(headers):
                ws.cell(row=r, column=c, value=val)


def build_player_level_sheet():
    """生成 Player Level 表：levelId, levelUpExp，共 120 行。"""
    headers = ["levelId", "levelUpExp"]
    types = ["int", "int"]
    comment = ["等级", "升到下一级所需经验"]
    rows = [[i, exp] for i, exp in enumerate(PLAYER_LEVEL_UP_EXP, start=1)]
    return comment, headers, types, rows


def build_weapons_sheet():
    """生成 Weapon Weapons 表。"""
    comment = ["唯一id", "游戏内展示名", "武器类型", "基础伤害", "伤害半径", "冷却", "攻击距离", "等级要求",
               "射线长度", "穿透块数", "爆炸尺寸", "引信时间", "钻头宽度",
               "层数", "半径", "公转速度", "自转速度", "子弹数", "反弹次数", "子弹速度"]
    rows = []
    for w in WEAPONS_DATA:
        row = []
        for h in WEAPONS_HEADERS:
            row.append(w.get(h))
        rows.append(row)
    return comment, WEAPONS_HEADERS, WEAPONS_TYPES, rows


def main():
    # -------- Player.xlsx --------
    wb_player = Workbook()
    ws_level = wb_player.active
    ws_level.title = "Level"
    comment, headers, types, rows = build_player_level_sheet()
    write_sheet(ws_level, comment, headers, types, rows)
    out_player = OUTPUT_DIR / "Player.xlsx"
    wb_player.save(out_player)
    print(f"Wrote: {out_player}")

    # -------- Weapon.xlsx --------
    wb_weapon = Workbook()
    ws_weapons = wb_weapon.active
    ws_weapons.title = "Weapons"
    comment, headers, types, rows = build_weapons_sheet()
    write_sheet(ws_weapons, comment, headers, types, rows)
    out_weapon = OUTPUT_DIR / "Weapon.xlsx"
    wb_weapon.save(out_weapon)
    print(f"Wrote: {out_weapon}")

    print("Done. Run ExcelToLua3.py to export to ExcelConfig.")


if __name__ == "__main__":
    main()
