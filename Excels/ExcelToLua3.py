"""
Excel to Lua table converter.

Excel format (single sheet, one row per record):
- Header row defines field keys (default row 1).
- Optional type row defines field types (e.g. string/int/float/bool/table).
- Optional display row (like Chinese hints) can be placed above header row.
- Nested fields use dot paths, e.g. loot.Stone, loot.Coal.
- Empty cells are ignored (field not written).
- Columns starting with # are ignored (notes for designers).

Key column:
- Default: column named "id" (case-insensitive), otherwise the first non-empty header.
- Use --key or --key-col to override.
- Key can be number or string; non-identifier keys are written as ["..."] or [123].

Type rules (auto or typed):
- number -> Lua number
- true/false (case-insensitive) -> Lua boolean
- nil (case-insensitive) -> Lua nil
- other text -> Lua string (escaped)
- table (cell text syntax):
  - key=value 优先；也支持 key:value。嵌套路径仍用 "a:b:c" -> { a = { b = c } }
  - "1|2|3" -> { 1, 2, 3 }
  - "itemId=10101,itemCount=3|itemId=10102,itemCount=5"
    -> { { itemId = 10101, itemCount = 3 }, { itemId = 10102, itemCount = 5 } }
  - 输出时：可整数化的 key 写为 [1] = value，不写 ["1"]

Usage:
  python ExcelToLua.py --input path.xlsx --output out.lua
  python ExcelToLua.py --input path.xlsx --output out.lua --sheet Sheet1
  python ExcelToLua.py --input path.xlsx --output out.lua --root-key Blocks
  python ExcelToLua.py --input path.xlsx --output out.lua --header-row 2 --type-row 3
  python ExcelToLua.py  (batch convert all .xlsx in current folder)
"""

from __future__ import annotations  # 允许在类型注解中使用 str | None 等 Python 3.10+ 语法

import argparse  # 命令行参数解析
import os
import re  # 正则表达式，用于标识符校验
import sys
from decimal import Decimal  # 高精度浮点数格式化，避免科学计数法
from pathlib import Path  # 跨平台路径处理
from typing import Any, Dict, Iterable, List, Tuple

try:
    from openpyxl import load_workbook  # openpyxl: 读取 .xlsx 文件的第三方库
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install with:\n"
        "  pip install -r requirements.txt\n"
        f"Details: {exc}"
    )

# ---------- 哨兵对象 ----------
NIL = object()   # 代表 Lua 的 nil 值（区分于 Python None / 空白单元格）
SKIP = object()  # 代表"跳过此字段不输出"（空单元格且不在 FORCE_NIL_FIELDS 中时使用）

# ---------- 常量 ----------
ID_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")  # 合法 Lua 标识符正则：字母/下划线开头
NUMBER_TYPES = (int, float)  # Python 数值类型元组，用于 isinstance 判断
KNOWN_TYPES = {"int", "integer", "float", "number", "string", "str", "bool", "boolean", "table"}  # 类型行可用的合法类型名
INDENT = "    "  # Lua 输出缩进：4 个空格

# ====================================================================
# 批处理模式全局配置（无 --input 参数时使用）
# ====================================================================
AUTO_DETECT_LAYOUT = True          # 是否自动探测表头/类型行/数据起始行的位置
DEFAULT_HEADER_ROW = 2             # 手动模式下，表头所在行（1-based）
DEFAULT_TYPE_ROW = 3               # 手动模式下，类型声明所在行
DEFAULT_DATA_START_ROW = 4         # 手动模式下，数据起始行
DEFAULT_KEY_NAME: str | None = None  # 按名称指定 Key 列（优先级低于 KEY_COL）
DEFAULT_KEY_COL: int | None = None   # 按列号指定 Key 列（1-based）

ROOT_KEY_MODE = "none"             # Lua 输出的根键模式：none=无根键 / sheet=Sheet名 / excel=文件名 / excel+sheet=文件名+Sheet名
ASSIGN_PATH: str | None = None    # 若设置，生成 `ASSIGN_PATH = {...}` 赋值语句（如 "RebirthPresets.RebirthConfig"）
FREEZE_RETURN = True               # 输出末尾是否追加 `return table.freeze(...)` 使 Luau 表只读
OUTPUT_DIR: Path | None = Path(
    "../src/ReplicatedStorage/ExcelConfig"
)  # 输出目录：None=与 xlsx 同目录；相对路径基于脚本所在目录(Excels)，导出到 ExcelConfig
OUTPUT_NAME_MODE = "excel+sheet"   # 输出文件名模式：sheet=Sheet名 / excel=文件名 / excel+sheet=文件名+Sheet名
SKIP_SHEETS = {"注释"}              # 批处理时跳过的 Sheet 名称集合
# 空单元格时仍输出 `field = nil`（而非省略该字段）的字段名集合
FORCE_NIL_FIELDS = {"key", "productid"}


# ====================================================================
# 辅助函数：规范化 / 类型推断
# ====================================================================

def _normalize_header(value: Any) -> str | None:
    """将表头单元格的值规范化为字符串；空白或 None 返回 None。"""
    if value is None:
        return None
    text = str(value).strip()  # 去除首尾空白
    if not text:
        return None
    return text


def _normalize_type(value: Any) -> str | None:
    """将类型行单元格的值规范化为小写字符串；空白或 None 返回 None。"""
    if value is None:
        return None
    text = str(value).strip().lower()  # 统一转小写以便匹配 KNOWN_TYPES
    return text or None


def _infer_scalar(text: str) -> Any:
    """
    从纯文本自动推断标量类型：
    - "true"/"false" -> Python bool
    - "nil"          -> NIL 哨兵
    - 含小数点      -> float
    - 纯数字        -> int
    - 其他          -> 原始字符串
    用于 table 类型单元格内部子值的解析。
    """
    raw = text.strip()
    if raw == "":
        return ""
    low = raw.lower()
    if low == "true":
        return True
    if low == "false":
        return False
    if low == "nil":
        return NIL
    try:
        if "." in raw:          # 含小数点 -> 浮点数
            return float(raw)
        return int(raw)         # 否则尝试整数
    except ValueError:
        return raw              # 无法转数值则保留字符串


# 正则：匹配 [N]=value 格式的方括号数字键，如 [1]=0.9
_BRACKET_KEY_RE = re.compile(r"^\[(\d+)\]\s*=\s*(.+)$")


def _parse_table_text(value: Any) -> Any:
    """
    解析 table 类型单元格的文本语法，返回 Python dict 或 list。
    
    支持的语法格式：
    1. "1|2|3"                             -> [1, 2, 3]（纯数组）
    2. "key=value"  或 "key:value"        -> { key: value }（键值对）
    3. "a:b:c"                             -> { a: { b: c } }（嵌套路径，冒号分隔）
    4. "itemId=10101,itemCount=3|itemId=10102,itemCount=5"
       -> [{ itemId: 10101, itemCount: 3 }, { itemId: 10102, itemCount: 5 }]
       （多对象数组，| 分隔条目，, 分隔键值对）
    5. 混合情况 -> 数组部分用数字索引 [1],[2]...，字典部分保留字符串键
    6. "{Stone=0.9002,Coal=0.0738}"        -> { Stone: 0.9002, Coal: 0.0738 }
       （花括号包裹格式，自动去除外层 {}）
    7. "{[1]=0.9,[2]=0.1}"                 -> { 1: 0.9, 2: 0.1 }
       （方括号数字键格式）
    8. 多行单元格内容（换行符视为逗号分隔）
    """
    text = str(value).strip()
    if text == "":
        return {}

    # --- 预处理：将换行符统一替换为逗号（支持多行单元格） ---
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\n", ",")

    # --- 预处理：去除外层花括号 {} ---
    # 花括号模式下，逗号是顶层分隔符（等同于无花括号时的 |）
    braced = False
    if text.startswith("{") and text.endswith("}"):
        text = text[1:-1].strip()           # 去掉最外层 { }
        braced = True
    if text == "":
        return {}

    root: Dict[str, Any] = {}   # 收集 key=value 或 key:nested:value 形式的键值对
    array: List[Any] = []       # 收集数组元素（无键的纯值 或 逗号分隔的对象）

    # 花括号模式：以逗号为顶层分隔符，每个 token 是一个 key=value
    # 非花括号模式：以 | 为顶层分隔符，逗号在条目内部表示多键值对的对象
    top_entries = text.split(",") if braced else text.split("|")

    for entry in top_entries:
        entry = entry.strip()
        if entry == "":
            continue

        # --- 情况 A（仅非花括号模式）：条目内含逗号 -> "key=val,key=val" 格式的对象 ---
        if not braced and "," in entry:
            pairs = [part.strip() for part in entry.split(",") if part.strip()]
            if not pairs:
                continue
            obj: Dict[str, Any] = {}
            valid_pairs = True
            for pair in pairs:
                bracket_match = _BRACKET_KEY_RE.match(pair)
                if bracket_match:
                    key = bracket_match.group(1)
                    val_str = bracket_match.group(2)
                    obj[key] = _infer_scalar(val_str)
                elif "=" in pair:
                    key, val = pair.split("=", 1)
                    obj[key.strip()] = _infer_scalar(val)
                elif ":" in pair:
                    key, val = pair.split(":", 1)
                    obj[key.strip()] = _infer_scalar(val)
                else:
                    valid_pairs = False
                    break
            if valid_pairs:
                array.append(obj)
            else:
                array.append(_infer_scalar(entry))
            continue

        # --- 情况 B'：支持 [N]=value 方括号数字键（如 [1]=0.9） ---
        bracket_match = _BRACKET_KEY_RE.match(entry)
        if bracket_match:
            key = bracket_match.group(1)       # 数字键字符串，如 "1"
            val_str = bracket_match.group(2)
            root[key] = _infer_scalar(val_str)
            continue

        # --- 情况 B：条目含 = 但不含 : -> 单个 key=value ---
        if "=" in entry and ":" not in entry:
            parts = [part.strip() for part in entry.split("=", 1)]
            if len(parts) == 2 and parts[0]:
                root[parts[0]] = _infer_scalar(parts[1])
                continue

        # --- 情况 C：条目含 : -> 嵌套路径 a:b:c -> { a = { b = c } } ---
        if ":" in entry:
            parts = [part.strip() for part in entry.split(":")]
            if len(parts) >= 2:
                *keys, val = parts          # 最后一个元素是值，前面都是键路径
                cur = root
                for key in keys[:-1]:       # 逐层创建嵌套字典
                    if key not in cur or not isinstance(cur[key], dict):
                        cur[key] = {}
                    cur = cur[key]
                last = keys[-1]
                cur[last] = _infer_scalar(val)
                continue

        # --- 情况 D：纯标量值，加入数组 ---
        array.append(_infer_scalar(entry))

    # --- 合并 root（字典部分）和 array（数组部分） ---
    if root and array:
        merged: Dict[Any, Any] = {}
        for idx, item in enumerate(array, start=1):  # 数组元素用 1-based 数字索引
            merged[idx] = item
        for key, val in root.items():
            merged[key] = val
        return merged
    if root:
        return root       # 仅有字典
    if array:
        return array      # 仅有数组
    return {}


def _parse_cell(value: Any, type_name: str | None = None) -> Any:
    """
    解析单个单元格的值，根据可选的类型声明进行强制转换。
    
    返回值：
    - SKIP  -> 空白单元格，调用者决定是否省略或输出 nil
    - NIL   -> 显式写 "nil" 的单元格
    - 其他  -> 转换后的 Python 值（bool / int / float / str / dict / list）
    """
    if value is None:                                # 空单元格
        return SKIP
    if isinstance(value, str) and value.strip() == "":  # 仅空白字符
        return SKIP

    type_name = _normalize_type(type_name)  # 规范化类型名（小写）

    # --- 有显式类型声明时，强制按类型转换 ---
    if type_name:
        if type_name in ("string", "str"):           # 强制字符串
            return str(value)
        if type_name in ("int", "integer"):          # 强制整数
            if isinstance(value, bool):
                return int(value)
            if isinstance(value, NUMBER_TYPES):
                return int(value)
            try:
                return int(str(value).strip())
            except ValueError:
                raise ValueError(f"Invalid int value: {value}")
        if type_name in ("float", "number"):         # 强制浮点数
            if isinstance(value, bool):
                return float(int(value))
            if isinstance(value, NUMBER_TYPES):
                return float(value)
            try:
                return float(str(value).strip())
            except ValueError:
                raise ValueError(f"Invalid float value: {value}")
        if type_name in ("bool", "boolean"):         # 强制布尔
            if isinstance(value, bool):
                return value
            text = str(value).strip().lower()
            if text == "true":
                return True
            if text == "false":
                return False
            raise ValueError(f"Invalid boolean value: {value}")
        if type_name == "table":                     # 强制 table（调用专用解析器）
            if isinstance(value, str) and value.strip().lower() == "nil":
                return NIL
            return _parse_table_text(value)

    # --- 无类型声明时，自动推断 ---
    if isinstance(value, bool):            # openpyxl 可能直接返回 bool
        return value
    if isinstance(value, NUMBER_TYPES):    # openpyxl 数值单元格直接返回 int/float
        return value
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return SKIP
        low = text.lower()
        if low == "true":
            return True
        if low == "false":
            return False
        if low == "nil":
            return NIL
        return text                        # 普通字符串
    return str(value)                      # 其余类型统一转字符串


# ====================================================================
# Lua 输出格式化函数
# ====================================================================

def _format_number(value: float | int) -> str:
    """将 Python 数值格式化为 Lua 数字字面量字符串。"""
    if isinstance(value, bool):          # bool 是 int 的子类，需要优先判断
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():           # 3.0 -> "3"，去除多余小数点
            return str(int(value))
        text = format(value, ".15g")     # 最多15位有效数字
        if "e" in text or "E" in text:   # 含科学计数法时，用 Decimal 转为定点表示
            dec = Decimal(str(value))
            text = format(dec, "f").rstrip("0").rstrip(".")
        return text
    return str(value)


def _escape_string(text: str) -> str:
    """转义字符串中的特殊字符，使其安全嵌入 Lua 双引号字符串。"""
    return (
        text.replace("\\", "\\\\")       # 反斜杠
        .replace("\n", "\\n")            # 换行
        .replace("\r", "\\r")            # 回车
        .replace("\t", "\\t")            # 制表符
        .replace('"', '\\"')             # 双引号
    )


def _lua_string(text: str) -> str:
    """将 Python 字符串包装为 Lua 双引号字符串字面量。"""
    return f'"{_escape_string(text)}"'


def _lua_key(key: Any) -> str:
    """
    将 Python 键格式化为 Lua 表键语法：
    - 数字键    -> [123]
    - 字符串 "3" 可整数化时 -> [3]（而非 ["3"]）
    - 合法标识符 -> fieldName（无需中括号）
    - 其他      -> ["some key"]
    """
    if isinstance(key, NUMBER_TYPES) and not isinstance(key, bool):
        return f"[{_format_number(key)}]"          # 数字键：[123]
    if isinstance(key, str):
        try:
            n = int(key)
            if str(n) == key:                      # "3" -> [3]
                return f"[{n}]"
        except ValueError:
            pass
        if ID_RE.match(key):                       # 合法标识符：直接写
            return key
    return f"[{_lua_string(str(key))}]"            # 其他：["key"]


def _to_lua(value: Any, indent: int = 0) -> str:
    """
    递归地将 Python 值序列化为 Lua 表达式字符串。
    支持 nil / bool / number / string / list(序列表) / dict(键值表)。
    """
    if value is NIL:
        return "nil"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return _format_number(value)
    if isinstance(value, str):
        return _lua_string(value)
    if isinstance(value, list):                    # 序列表：{ val1, val2, ... }
        if not value:
            return "{}"
        indent_str = INDENT * indent
        child_indent = INDENT * (indent + 1)
        lines = ["{"]
        for child in value:
            child_text = _to_lua(child, indent + 1)
            lines.append(f"{child_indent}{child_text},")
        lines.append(f"{indent_str}}}")
        return "\n".join(lines)
    if isinstance(value, dict):                    # 键值表：{ key = val, ... }
        if not value:
            return "{}"
        indent_str = INDENT * indent
        child_indent = INDENT * (indent + 1)
        lines: List[str] = ["{"]
        for key, child in value.items():
            child_text = _to_lua(child, indent + 1)
            lines.append(f"{child_indent}{_lua_key(key)} = {child_text},")
        lines.append(f"{indent_str}}}")
        return "\n".join(lines)
    return _lua_string(str(value))                 # 兜底：转字符串


# ====================================================================
# Excel 行/列解析函数
# ====================================================================

def _row_has_data(values: Iterable[Any]) -> bool:
    """判断一行是否至少有一个非空单元格（用于跳过空行）。"""
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return True
    return False


def _parse_headers(
    header_row: Iterable[Any],
    type_row: Iterable[Any] | None,
) -> Tuple[List[Tuple[str, ...] | None], List[str | None], Dict[str, int]]:
    """
    解析表头行和类型行，返回三个结构：
    - specs:      每列的字段路径元组（如 ("loot","Stone")），跳过的列为 None
    - types:      每列的类型声明字符串，无声明为 None
    - header_map: 表头名称 -> 列索引的映射字典
    """
    specs: List[Tuple[str, ...] | None] = []  # 每列的字段路径
    types: List[str | None] = []              # 每列的类型声明
    header_map: Dict[str, int] = {}           # 表头名 -> 列索引
    for idx, raw in enumerate(header_row):
        header = _normalize_header(raw)
        if header is None:                    # 空表头 -> 跳过该列
            specs.append(None)
            types.append(None)
            continue
        if header.startswith("#"):            # # 开头 -> 设计师注释列，跳过
            specs.append(None)
            types.append(None)
            continue
        parts = [part.strip() for part in header.split(".")]  # 支持点号嵌套路径：loot.Stone
        if any(part == "" for part in parts):
            raise ValueError(f"Invalid header '{header}' (empty path segment).")
        specs.append(tuple(parts))
        header_map[header] = idx
        if type_row is not None and idx < len(type_row):  # 对应列的类型声明
            types.append(_normalize_type(type_row[idx]))
        else:
            types.append(None)
    return specs, types, header_map


def _set_path(
    target: Dict[str, Any],
    path: Tuple[str, ...],
    value: Any,
    row_index: int,
    header_name: str,
) -> None:
    """
    将值写入嵌套字典的指定路径。
    如 path=("loot","Stone") 时：target["loot"]["Stone"] = value
    自动创建中间层级的字典；冲突时抛出错误。
    """
    current = target
    for key in path[:-1]:                     # 遍历中间路径，逐层深入/创建字典
        if key not in current:
            current[key] = {}
        elif not isinstance(current[key], dict):
            raise ValueError(
                f"Row {row_index}: field '{header_name}' conflicts with "
                f"existing non-table value at '{key}'."
            )
        current = current[key]
    last = path[-1]                           # 路径的最后一段，赋值
    if last in current and isinstance(current[last], dict):
        raise ValueError(
            f"Row {row_index}: field '{header_name}' conflicts with "
            f"existing table value at '{last}'."
        )
    current[last] = value


def _get_row_values(worksheet, row_index: int) -> Tuple[Any, ...]:
    """从 openpyxl worksheet 中读取指定行的所有单元格值（1-based 行号）。"""
    row = next(
        worksheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True),
        None,
    )
    return row or ()


def _row_is_types(row: Iterable[Any]) -> bool:
    """判断一行是否全部由合法类型名组成（用于自动探测类型行）。"""
    values = [v for v in row if v is not None and str(v).strip() != ""]
    if not values:
        return False
    for v in values:
        t = _normalize_type(v)
        if not t or t not in KNOWN_TYPES:    # 存在非类型名 -> 不是类型行
            return False
    return True


def _row_has_headers(row: Iterable[Any]) -> bool:
    """判断一行是否包含表头（有非空值且不全是类型名）。"""
    values = [v for v in row if v is not None and str(v).strip() != ""]
    if not values:
        return False
    return not _row_is_types(values)         # 不是类型行 -> 是表头行


def _resolve_key_column(
    header_map: Dict[str, int],
    specs: List[Tuple[str, ...] | None],
    key_name: str | None,
    key_col: int | None,
) -> int:
    """
    确定哪一列作为 Lua 表的主键（即输出 { [key] = record } 的 key）。
    优先级：--key-col > --key > 名为 "id" 的列 > 第一个非空表头列。
    返回 0-based 列索引。
    """
    if key_col is not None:                  # 用户显式指定列号（1-based）
        if key_col < 1:
            raise ValueError("key column index must be 1-based.")
        return key_col - 1
    if key_name:                             # 用户显式指定列名
        if key_name not in header_map:
            raise ValueError(f"Key column '{key_name}' not found in header.")
        return header_map[key_name]
    for name, idx in header_map.items():     # 自动查找名为 "id" 的列（忽略大小写）
        if name.lower() == "id":
            return idx
    for idx, spec in enumerate(specs):       # 兜底：第一个有效列
        if spec is not None:
            return idx
    raise ValueError("No usable column found for key.")


def build_table_from_worksheet(
    worksheet,
    header_row_index: int = 1,
    type_row_index: int | None = None,
    data_start_row: int | None = None,
    key_name: str | None = None,
    key_col: int | None = None,
) -> Dict[Any, Dict[str, Any]]:
    """
    核心函数：将一个 worksheet 转换为 { key: record_dict } 的 Python 字典。
    
    流程：
    1. 读取表头行 -> 解析字段路径（specs）和类型声明（types）
    2. 确定主键列
    3. 逐行遍历数据区域，解析每个单元格，组装记录字典
    """
    header_row = _get_row_values(worksheet, header_row_index)  # 读取表头行
    if not header_row:
        raise ValueError("Worksheet is empty or missing header row.")

    type_row = _get_row_values(worksheet, type_row_index) if type_row_index else None  # 读取类型行（可选）
    specs, types, header_map = _parse_headers(header_row, type_row)  # 解析表头和类型

    # 构建表头名列表（用于错误信息）
    header_names: List[str] = []
    for spec in specs:
        if spec is None:
            header_names.append("")
        else:
            header_names.append(".".join(spec))

    key_col_index = _resolve_key_column(header_map, specs, key_name, key_col)  # 确定主键列

    # 确定数据起始行（未指定时自动推算）
    if data_start_row is None:
        if type_row_index:
            data_start_row = type_row_index + 1    # 类型行下一行
        else:
            data_start_row = header_row_index + 1  # 表头下一行

    # --- 逐行读取数据，构建输出表 ---
    table: Dict[Any, Dict[str, Any]] = {}
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True),
        start=data_start_row,
    ):
        if not _row_has_data(row):               # 跳过空行
            continue

        # -- 提取并校验主键 --
        key_value = row[key_col_index] if key_col_index < len(row) else None
        if key_value is None or (isinstance(key_value, str) and key_value.strip() == ""):
            raise ValueError(f"Row {row_index}: key is required.")

        key_type = types[key_col_index] if key_col_index < len(types) else None
        parsed_key = _parse_cell(key_value, key_type)  # 按类型解析主键
        if parsed_key is SKIP or parsed_key is NIL:
            raise ValueError(f"Row {row_index}: key is required.")
        if isinstance(parsed_key, str):
            parsed_key = parsed_key.strip()
            if parsed_key == "":
                raise ValueError(f"Row {row_index}: key is required.")
        if parsed_key in table:                  # 主键不能重复
            raise ValueError(f"Row {row_index}: duplicate key '{parsed_key}'.")

        # -- 逐列解析，组装记录 --
        record: Dict[str, Any] = {}
        for col_index, spec in enumerate(specs):
            if spec is None:                     # 跳过无效列（空表头 / # 注释列）
                continue
            cell_value = row[col_index] if col_index < len(row) else None
            type_name = types[col_index] if col_index < len(types) else None
            parsed = _parse_cell(cell_value, type_name)
            if parsed is SKIP:                   # 空单元格 -> 输出为 nil（保留字段）
                parsed = NIL
            header_name = header_names[col_index]
            _set_path(record, spec, parsed, row_index, header_name)  # 写入嵌套路径

        table[parsed_key] = record               # 以主键为 key 存入总表

    return table


def _detect_layout(worksheet) -> Tuple[int, int | None, int]:
    """
    自动探测 Excel 的布局结构，返回 (表头行, 类型行, 数据起始行)。
    
    探测策略（逐行检查前3行）：
    - 第1行类型、第2行显示名、第3行表头 -> (3, 1, 4)  ← 类型在最上面
    - 第1行显示名、第2行表头、第3行类型 -> (2, 3, 4)  ← 类型在最下面
    - 第1行表头、第2行类型              -> (1, 2, 3)
    - 第1行表头、无类型行               -> (1, None, 2)
    """
    row1 = _get_row_values(worksheet, 1)  # 读取前三行用于判断
    row2 = _get_row_values(worksheet, 2)
    row3 = _get_row_values(worksheet, 3)

    # 类型行在第1行，表头在第3行（第2行为中文显示名）
    if _row_is_types(row1) and _row_has_headers(row3):
        return 3, 1, 4
    # 类型行在第3行，表头在第2行（第1行为中文显示名）
    if _row_has_headers(row2) and _row_is_types(row3):
        return 2, 3, 4
    # 表头在第1行，类型在第2行
    if _row_has_headers(row1) and _row_is_types(row2):
        return 1, 2, 3
    return 1, None, 2                                   # 第1行表头，无类型行


def convert_excel_to_lua(
    input_path: Path,
    sheet_name: str | None,
    root_key: str | None,
    var_name: str | None,
    header_row: int,
    type_row: int | None,
    data_start_row: int | None,
    key_name: str | None,
    key_col: int | None,
) -> str:
    """
    单文件模式（--input）：读取一个 Excel 文件的指定 Sheet，返回 Lua 源码字符串。
    
    输出格式取决于参数组合：
    - var_name   -> `local VarName = {...}; return table.freeze(VarName)`
    - ASSIGN_PATH -> `Path.To.Table = {...}; return table.freeze(Path.To.Table)`
    - 默认       -> `return table.freeze({...})`
    """
    workbook = load_workbook(input_path, data_only=True)  # data_only=True：读取公式的计算结果
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]   # 打开指定 Sheet
    else:
        worksheet = workbook.active        # 默认使用活动 Sheet

    # 将 worksheet 转换为 Python 字典
    table = build_table_from_worksheet(
        worksheet,
        header_row_index=header_row,
        type_row_index=type_row,
        data_start_row=data_start_row,
        key_name=key_name,
        key_col=key_col,
    )
    if root_key:                           # 可选：用 root_key 包裹一层
        output_table: Dict[str, Any] = {root_key: table}
    else:
        output_table = table

    # --- 根据输出模式生成 Lua 代码 ---
    if var_name:
        # Luau 模块风格：local VarName = {...}\nreturn table.freeze(VarName)
        text = f"local {var_name} = {_to_lua(output_table)}\n"
        if FREEZE_RETURN:
            text += f"return table.freeze({var_name})\n"
        else:
            text += f"return {var_name}\n"
        return text

    if ASSIGN_PATH:                        # 赋值路径模式
        text = f"{ASSIGN_PATH} = {_to_lua(output_table)}\n"
        if FREEZE_RETURN:
            text += f"return table.freeze({ASSIGN_PATH})\n"
        return text
    if FREEZE_RETURN:                      # 默认：直接 return frozen table
        return f"return table.freeze({_to_lua(output_table)})\n"
    return f"return {_to_lua(output_table)}\n"


def _resolve_root_key(excel_name: str, sheet_name: str) -> str | None:
    """根据 ROOT_KEY_MODE 计算 Lua 输出的根键名称。"""
    mode = ROOT_KEY_MODE.lower()
    if mode == "sheet":                    # 仅 Sheet 名
        return sheet_name
    if mode == "excel":                    # 仅 Excel 文件名（不含扩展名）
        return excel_name
    if mode == "excel+sheet":              # 文件名 + Sheet 名拼接
        return f"{excel_name}{sheet_name}"
    return None                            # none -> 不包裹根键


def _resolve_output_name(excel_name: str, sheet_name: str) -> str:
    """根据 OUTPUT_NAME_MODE 计算输出 .lua 文件的基础名。"""
    mode = OUTPUT_NAME_MODE.lower()
    if mode == "sheet":
        return sheet_name
    if mode == "excel":
        return excel_name
    return f"{excel_name}{sheet_name}"     # 默认 excel+sheet


def _resolve_output_dir(input_path: Path) -> Path:
    """
    确定输出目录的绝对路径。
    - OUTPUT_DIR 为 None  -> 与输入文件同目录
    - OUTPUT_DIR 是绝对路径 -> 直接使用
    - OUTPUT_DIR 是相对路径 -> 相对于本脚本所在目录（Excels/）解析
    """
    if OUTPUT_DIR is None:
        return input_path.parent
    if OUTPUT_DIR.is_absolute():
        return OUTPUT_DIR
    return (Path(__file__).resolve().parent / OUTPUT_DIR).resolve()


def _convert_workbook_batch(input_path: Path) -> None:
    """
    批处理模式：将一个 .xlsx 文件中的每个 Sheet 分别导出为独立的 .lua 文件。
    跳过 SKIP_SHEETS 中列出的 Sheet。
    """
    workbook = load_workbook(input_path, data_only=True)  # 打开工作簿
    excel_name = input_path.stem                          # 文件名（无扩展名）
    output_dir = _resolve_output_dir(input_path)          # 计算输出目录
    output_dir.mkdir(parents=True, exist_ok=True)         # 确保输出目录存在

    for sheet_name in workbook.sheetnames:                # 遍历所有 Sheet
        if sheet_name in SKIP_SHEETS:                     # 跳过注释等特殊 Sheet
            continue
        worksheet = workbook[sheet_name]

        # 自动探测或使用默认布局
        if AUTO_DETECT_LAYOUT:
            header_row, type_row, data_start_row = _detect_layout(worksheet)
        else:
            header_row = DEFAULT_HEADER_ROW
            type_row = DEFAULT_TYPE_ROW
            data_start_row = DEFAULT_DATA_START_ROW

        root_key = _resolve_root_key(excel_name, sheet_name)  # 计算根键
        table = build_table_from_worksheet(                   # 解析 worksheet -> Python dict
            worksheet,
            header_row_index=header_row,
            type_row_index=type_row,
            data_start_row=data_start_row,
            key_name=DEFAULT_KEY_NAME,
            key_col=DEFAULT_KEY_COL,
        )
        output_table: Dict[str, Any] = {root_key: table} if root_key else table  # 可选根键包裹

        # 生成 Lua 代码文本
        if ASSIGN_PATH:
            lua_text = f"{ASSIGN_PATH} = {_to_lua(output_table)}\n"
            if FREEZE_RETURN:
                lua_text += f"return table.freeze({ASSIGN_PATH})\n"
        else:
            lua_text = (
                f"return table.freeze({_to_lua(output_table)})\n"
                if FREEZE_RETURN
                else f"return {_to_lua(output_table)}\n"
            )

        # 写入 .lua 文件
        output_name = _resolve_output_name(excel_name, sheet_name)
        output_path = output_dir / f"{output_name}.lua"
        output_path.write_text(lua_text, encoding="utf-8")
        print(f"Wrote Lua table to: {output_path}")


# ====================================================================
# 命令行参数解析
# ====================================================================

def _parse_args(argv: List[str]) -> argparse.Namespace:
    """定义并解析所有命令行参数。"""
    parser = argparse.ArgumentParser(description="Convert Excel to Lua table.")
    parser.add_argument("--input", default=None, help="输入 .xlsx 文件路径（省略则进入批处理模式）")
    parser.add_argument("--output", default=None, help="输出 .lua 文件路径（省略则自动生成）")
    parser.add_argument("--sheet", default=None, help="指定 Sheet 名称（省略则使用活动 Sheet）")
    parser.add_argument(
        "--var",
        "--var-name",
        dest="var_name",
        default=None,
        help="生成 `local <var> = {...}` 风格的 Luau 模块",
    )
    parser.add_argument(
        "--root-key",
        default=None,
        help="用根键包裹输出：return { rootKey = {...} }",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="表头行号（1-based），默认 1",
    )
    parser.add_argument(
        "--type-row",
        type=int,
        default=None,
        help="类型行号（1-based），可选",
    )
    parser.add_argument(
        "--data-start-row",
        type=int,
        default=None,
        help="数据起始行号（1-based），可选",
    )
    parser.add_argument(
        "--key",
        default=None,
        help="主键列名（表头行中的字段名）",
    )
    parser.add_argument(
        "--key-col",
        type=int,
        default=None,
        help="主键列号（1-based）",
    )
    return parser.parse_args(argv)


# ====================================================================
# 主入口
# ====================================================================

def main(argv: List[str]) -> int:
    """
    程序主入口，返回 exit code（0=成功，1=转换错误，2=文件未找到）。
    
    两种运行模式：
    - 无 --input 参数 -> 批处理模式：扫描脚本所在目录所有 .xlsx，逐个转换
    - 有 --input 参数 -> 单文件模式：转换指定的 .xlsx
    """
    args = _parse_args(argv)

    if not args.input:
        # ===== 批处理模式 =====
        current_dir = Path(__file__).resolve().parent  # 脚本所在目录（Excels/）
        xlsx_files = [
            p
            for p in current_dir.iterdir()
            if p.is_file()
            and p.suffix.lower() == ".xlsx"
            and not p.name.startswith("~$")            # 排除 Excel 临时锁定文件
        ]
        if not xlsx_files:
            print("No .xlsx files found in current folder.", file=sys.stderr)
            return 2
        for file_path in xlsx_files:
            try:
                _convert_workbook_batch(file_path)     # 逐个工作簿批量转换
            except Exception as exc:
                print(f"Conversion failed for {file_path}: {exc}", file=sys.stderr)
                return 1
        return 0

    # ===== 单文件模式 =====
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 2

    # 确定输出路径
    output_path = Path(args.output) if args.output else None
    if output_path is None:
        output_dir = _resolve_output_dir(input_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{input_path.stem}.lua"

    # 执行转换
    try:
        lua_text = convert_excel_to_lua(
            input_path,
            sheet_name=args.sheet,
            root_key=args.root_key,
            var_name=args.var_name,
            header_row=args.header_row,
            type_row=args.type_row,
            data_start_row=args.data_start_row,
            key_name=args.key,
            key_col=args.key_col,
        )
    except Exception as exc:
        print(f"Conversion failed: {exc}", file=sys.stderr)
        return 1

    # 写入输出文件
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(lua_text, encoding="utf-8")
    print(f"Wrote Lua table to: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))  # 将 main 的返回值作为进程退出码
